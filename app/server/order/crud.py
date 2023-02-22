import os
from typing import List

from fastapi import HTTPException, Response
from openpyxl.styles import Alignment
from sqlalchemy import and_
from sqlalchemy.orm import Session, aliased
from datetime import datetime

from app.models.domain.Error_handler import UnicornException
from app.models.domain.order import Order
from app.models.domain.order_issue import OrderIssue
from app.models.domain.user import User
from app.models.domain.user_mark_order import UserMarkOrder
from app.models.schemas.order import OrderModifyModel, OrderViewModel, OrderCreateModel, OrderMarkPost

# import jpype
# import asposecells

from app.server.authentication import AuthorityLevel

# jpype.startJVM()


def get_order_engineer_id(db: Session, order_id: int):
    order_db = db.query(Order).filter(Order.id == order_id).first()
    return order_db.engineer_id


def get_order_status(db: Session, order_id: int):
    order_db = db.query(Order).filter(Order.id == order_id).first()
    return order_db.status


def create_order(db: Session, reporter_id, company_name, order_create: OrderCreateModel):
    if not db.query(User).filter(User.id == order_create.client_id).first():
        raise UnicornException(name=create_order.__name__, description='client user not found', status_code=404)
    try:
        order_create.detail = str(order_create.detail)
        db_order = Order(**order_create.dict(),
                         reporter_id=reporter_id,
                         company_name=company_name)
        db.add(db_order)
        db.commit()
        db.refresh(db_order)
        db_order.detail = eval(db_order.detail)  # 因sqlite不能用Array存，所以先轉str，再轉list輸出
        db_order.file_name = eval(db_order.file_name)  # 因sqlite不能用Array存，所以先轉str，再轉list輸出
    except Exception as e:
        db.rollback()
        print(str(e))
        raise UnicornException(name=create_order.__name__, description=str(e), status_code=500)
    return db_order



def get_order_view_model(each_order, engineer_name, issue_name, mark):
    engineer_name = engineer_name if engineer_name else "未指派"
    issue_name = issue_name if issue_name else "未選擇問題種類"
    mark = mark if mark else False
    try:
        dir_path = os.getcwd() + f"/db_image/order_pic_name_by_id/{each_order.id}"
        all_file_name = os.listdir(dir_path)
    except:
        all_file_name = []
    return OrderViewModel(
        id=each_order.id,
        reporter_id=each_order.reporter_id,
        company_name=each_order.company_name,
        description=each_order.description,
        detail=eval(each_order.detail),
        engineer_name=engineer_name,
        issue_name=issue_name,
        mark=mark,
        status=each_order.status,
        created_at=each_order.created_at,
        updated_at=each_order.updated_at,
        file_name=all_file_name
    )


def get_all_order(db: Session, level, user_id, start_time, end_time):
    order_db = db.query(Order, User.name, OrderIssue.name, UserMarkOrder.mark).outerjoin(
        User, Order.engineer_id == User.id
    ).outerjoin(
        OrderIssue, Order.order_issue_id == OrderIssue.id
    ).outerjoin(
        UserMarkOrder, and_(
            UserMarkOrder.order_id == Order.id,
            UserMarkOrder.user_id == user_id
        )
    )
    # engineer only get h
    if level == 2:
        order_db = order_db.filter(Order.engineer_id.in_([0, user_id]))

    if level == 3:
        order_db = order_db.filter(Order.client_id == user_id)

    if start_time:
        order_db = order_db.filter(Order.created_at > start_time)

    if end_time:
        order_db = order_db.filter(Order.created_at < end_time)


    view_models = [get_order_view_model(each_order, engineer_name, issue_name, mark)
                   for each_order, engineer_name, issue_name, mark in order_db]
    return view_models


def get_some_order(db: Session, user_id, client_id_list, engineer_id_list, order_issue_id_list, status_list,
                   start_time, end_time):
    # Join the Order table with the User and OrderIssue tables
    order_db = db.query(Order, User.name, OrderIssue.name, UserMarkOrder.mark).outerjoin(
        User, Order.engineer_id == User.id
    ).outerjoin(
        OrderIssue, Order.order_issue_id == OrderIssue.id
    ).outerjoin(
        UserMarkOrder, and_(
            UserMarkOrder.order_id == Order.id,
            UserMarkOrder.user_id == user_id
        )
    )

    if client_id_list:
        order_db = order_db.filter(Order.client_id.in_(client_id_list))
    if engineer_id_list:
        order_db = order_db.filter(Order.engineer_id.in_(engineer_id_list))
    if order_issue_id_list:
        order_db = order_db.filter(Order.order_issue_id.in_(order_issue_id_list))
    if status_list:
        order_db = order_db.filter(Order.status.in_(status_list))
    if start_time:
        order_db = order_db.filter(Order.created_at > start_time)
    if end_time:
        order_db = order_db.filter(Order.created_at < end_time)

    # Convert each order to a view model
    view_models = [get_order_view_model(each_order, engineer_name, issue_name, mark)
                   for each_order, engineer_name, issue_name, mark in order_db]

    return view_models


def get_a_order(db, order_id, user_id):
    # Join the Order table with the User and OrderIssue tables
    order_db = db.query(Order, User.name, OrderIssue.name, UserMarkOrder.mark).outerjoin(
        User, Order.engineer_id == User.id
    ).outerjoin(
        OrderIssue, Order.order_issue_id == OrderIssue.id
    ).outerjoin(
        UserMarkOrder, and_(
            UserMarkOrder.order_id == Order.id,
            UserMarkOrder.user_id == user_id
        )
    ).filter(Order.id == order_id)

    view_models = [get_order_view_model(each_order, engineer_name, issue_name, mark)
                   for each_order, engineer_name, issue_name, mark in order_db]

    return view_models[0]


def get_a_order(db, order_id, user_id):
    order_db = db.query(Order, User.name, OrderIssue.name, UserMarkOrder.mark).outerjoin(
        User, Order.engineer_id == User.id
    ).outerjoin(
        OrderIssue, Order.order_issue_id == OrderIssue.id
    ).outerjoin(
        UserMarkOrder, and_(
            UserMarkOrder.order_id == Order.id,
            UserMarkOrder.user_id == user_id
        )
    )
    order_db = order_db.filter(Order.id == order_id)
    view_model = [get_order_view_model(each_order, engineer_name, issue_name, mark)
                  for each_order, engineer_name, issue_name, mark in order_db]

    return view_model[0]


def check_order_status(db: Session, order_id_list):
    order_db = db.query(Order).filter(Order.id.in_(order_id_list), Order.status != 0).first()
    if order_db:
        status = order_db.status
    else:
        status = None
    return status


def delete_order_by_id(db: Session, order_id_list):
    try:
        # delete all message in order
        from app.models.domain.order_message import OrderMessage
        db.query(OrderMessage).filter(OrderMessage.order_id.in_(order_id_list)).delete(synchronize_session=False)

        # delete all mark in order
        from app.models.domain.user_mark_order import UserMarkOrder
        db.query(UserMarkOrder).filter(UserMarkOrder.order_id.in_(order_id_list)).delete(synchronize_session=False)

        db.query(Order).filter(Order.id.in_(order_id_list)).delete(synchronize_session=False)
        db.commit()
    except Exception as e:
        db.rollback()
        raise UnicornException(name=delete_order_by_id.__name__, description=str(e), status_code=500)
    # 刪除圖片
    for order_id in order_id_list:
        file_path = os.getcwd() + "/db_image/order_pic_name_by_id/" + str(order_id) + ".jpg"
        if os.path.exists(file_path):
            os.remove(file_path)
    return "Order deleted successfully"


def modify_order_by_id(db: Session, order_modify: OrderModifyModel):
    order_db = db.query(Order).filter(Order.id == order_modify.order_id).first()
    if not order_db:
        raise UnicornException(name=modify_order_by_id.__name__, description='order not found', status_code=404)
    order_db.order_issue_id = order_modify.order_issue_id
    order_db.description = order_modify.description
    order_db.detail = str(order_modify.detail)
    order_db.updated_at = datetime.now()
    db.commit()
    return "Order modify successfully"


def check_modify_status_permission(db: Session, current_user, now_status: int, status: int, order_id: int):
    # check client change status auth
    if current_user.level == AuthorityLevel.client.value:  # client
        if not db.query(Order).filter(Order.id == order_id, Order.client_id == current_user.id).first():
            raise HTTPException(
                status_code=401,
                detail="order isn't yours"
            )
        if now_status != 2:
            raise HTTPException(
                status_code=401,
                detail="client only can change status when order finish"
            )
        if status not in [1, 3] and status != 2:
            raise HTTPException(
                status_code=401,
                detail="client only can change status to 'in process' or 'closing order' when order finish"
            )

    # check engineer change status auth
    elif current_user.level == AuthorityLevel.engineer.value:  # engineer
        if status in [3, 0]:
            raise HTTPException(
                status_code=401,
                detail="engineer only can change status to not appoint or finish"
            )

    # check pm change status auth
    elif current_user.level == AuthorityLevel.pm.value:  # pm
        if status == 3:
            raise HTTPException(
                status_code=401,
                detail="pm can't change status to close"
            )


def modify_order_status_by_id(db: Session, order_id: int, status: int):
    order_db = db.query(Order).filter(Order.id == order_id).first()
    if not order_db:
        raise UnicornException(
            name=modify_order_principal_engineer_by_id.__name__,
            description='order not found',
            status_code=404
        )
    # status
    order_db.status = status
    order_db.updated = datetime.now()
    db.commit()
    return order_db


def modify_order_principal_engineer_by_id(db: Session, order_id: int, engineer_id: int):
    order_db = db.query(Order).filter(Order.id == order_id, Order.id == order_id).first()
    if not order_db:
        raise UnicornException(
            name=modify_order_principal_engineer_by_id.__name__,
            description='order not found',
            status_code=404)

    status = 1 if order_db.status == 0 else order_db.status
    order_db.engineer_id = engineer_id
    order_db.status = status
    order_db.updated = datetime.now()
    db.commit()
    return order_db


def order_mark_by_user(db: Session, user_id, order_mark: OrderMarkPost):

    # Check if the user exists
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise UnicornException(
            name=order_mark_by_user.__name__,
            description='User does not exist',
            status_code=404)

    # Check if the order exists
    order = db.query(Order).filter(Order.id == order_mark.order_id).first()
    if not order:
        raise UnicornException(
            name=order_mark_by_user.__name__,
            description='Order does not exist',
            status_code=404)

    # Check if the user has marked the order
    mark_db = db.query(UserMarkOrder).filter(
            UserMarkOrder.user_id == user_id,
            UserMarkOrder.order_id == order_mark.order_id
        ).first()

    try:
        # If the mark is empty, delete the mark from the database
        if not order_mark.mark:
            if mark_db:
                db.delete(mark_db)
                db.commit()

        # If the mark is not empty, and the mark does not exist in the database, add it to the database
        else:
            if not mark_db:
                mark = UserMarkOrder(**order_mark.dict(), user_id=user_id)
                db.add(mark)
                db.commit()
                db.refresh(mark)
    except Exception as e:
        print(str(e))
        raise UnicornException(name=delete_order_by_id.__name__, description=str(e), status_code=500)

    return "Done"


async def upload_picture_to_folder(db: Session, order_id: int, client_id: int, picture_file):
    # order = db.query(Order).filter(Order.id == order_id, Order.client_id == client_id).first()
    # if not order:
    #     HTTPException(status_code=401, detail="client only can upload file to his order")

    file_dir = os.path.join(os.getcwd(), f"db_image/order_pic_name_by_id/{order_id}/")
    os.makedirs(file_dir, exist_ok=True)
    try:
        file_path = os.path.join(file_dir, picture_file.filename)
        with open(file_path, "wb") as f:
            f.write(await picture_file.read())
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail="Failed to upload image")
    return {"message": "Image uploaded successfully"}


def download_picture_from_folder(order_id, file_name):
    print("abc")
    file_path = os.getcwd() + f"/db_image/order_pic_name_by_id/{order_id}/" + file_name
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Image doesn't exist.")
    try:
        from starlette.responses import FileResponse

        # 如果檔案是圖片，直接顯示
        if file_name.endswith(".jpg") or file_name.endswith(".png"):
            image_path = os.getcwd() + f"/db_image/order_pic_name_by_id/{order_id}/{file_name}"
            return FileResponse(image_path, media_type="image/png")

        # 如果檔案是其他類型，下載
        else:
            file_path = os.getcwd() + f"/db_image/order_pic_name_by_id/{order_id}/{file_name}"
            return FileResponse(file_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to download image")


def delete_picture_from_folder(db: Session, order_id, file_name):
    file_path = os.getcwd() + f"/db_image/order_pic_name_by_id/{order_id}/" + str(file_name)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Image doesn't exist.")
    try:
        os.remove(file_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to Delete image")
    return "Image deleted successfully"


async def get_report(db, user_id):
    import openpyxl
    from asposecells.api import Workbook, FileFormatType, PdfSaveOptions
    from openpyxl.styles.alignment import Alignment
    from openpyxl.styles.borders import Border, Side

    # Create aliases for joined tables
    engineer_alias = aliased(User)
    reporter_alias = aliased(User)

    # Specify the primary table for the query using .select_from()
    order_db = db.query(Order, engineer_alias.name, reporter_alias.name, OrderIssue.name,
                        UserMarkOrder.mark).select_from(Order)

    # Join with the aliased tables
    order_db = order_db.outerjoin(
        engineer_alias, Order.engineer_id == engineer_alias.id
    ).outerjoin(
        reporter_alias, Order.reporter_id == reporter_alias.id
    ).outerjoin(
        OrderIssue, Order.order_issue_id == OrderIssue.id
    ).outerjoin(
        UserMarkOrder, and_(
            UserMarkOrder.order_id == Order.id,
            UserMarkOrder.user_id == user_id
        )
    )


    # Create directory for report files if it doesn't exist
    report_dir = os.path.join(os.getcwd(), "db_image/order/report/")
    os.makedirs(report_dir, exist_ok=True)

    # Load the Excel workbook
    workbook = openpyxl.load_workbook('order.xlsx')

    # Select the worksheet to edit
    worksheet = workbook['EFAY - 2023 JAN']

    # Create a list of weekday names to use in the report
    week_name = ["一", "二", "三", "四", "五", "六", "日"]

    # Set alignment for the cells
    align = Alignment(horizontal='center', vertical='center')

    # Set border
    thin_border = Border(bottom=Side(style='thin'))

    # 初始化上一行日期
    prev_date = None
    start_row = 3

    # Iterate over the query results and populate the worksheet
    for idx, (order, engineer_name, reporter_name, issue_name, mark) in enumerate(order_db, start=4):
        curr_date = order.created_at.date()

        # Merge cells with the same date
        if curr_date != prev_date:
            # Get the end row for the merged cells
            end_row = idx - 1

            # Merge cells and write the date
            # column B
            worksheet.merge_cells(
                start_row=start_row,
                start_column=2,
                end_row=end_row,
                end_column=2
            )

            # Merge cells and write the day of the week
            # column C
            worksheet.merge_cells(
                start_row=start_row,
                start_column=3,
                end_row=end_row,
                end_column=3
            )

            # Update the start row for the next merged cells
            start_row = idx

            # Update the previous date
            prev_date = curr_date

            # align
            worksheet[f"B{start_row}"].alignment = align
            worksheet[f"C{start_row}"].alignment = align

            # write line
            worksheet[f"B{start_row}"].border = thin_border
            worksheet[f"C{start_row}"].border = thin_border


        # Check if this is the last row
        if idx == len(list(order_db)) + 3:
            end_row = idx

            # column B
            worksheet.merge_cells(
                start_row=start_row,
                start_column=2,
                end_row=end_row,
                end_column=2
            )

            # column C
            worksheet.merge_cells(
                start_row=start_row,
                start_column=3,
                end_row=end_row,
                end_column=3
            )

        # Get the weekday name from the order date
        weekday_name = week_name[order.created_at.weekday()]

        # write the data into the merge
        worksheet.cell(row=start_row, column=2, value=prev_date)
        worksheet.cell(row=start_row, column=3, value=weekday_name)



        worksheet[f"D{idx}"] = reporter_name
        worksheet[f"D{idx}"].alignment = align
        worksheet[f"D{idx}"].border = thin_border

        worksheet[f"E{idx}"] = issue_name
        worksheet[f"E{idx}"].alignment = align
        worksheet[f"E{idx}"].border = thin_border

        worksheet[f"F{idx}"] = "1"
        worksheet[f"F{idx}"].alignment = align
        worksheet[f"F{idx}"].border = thin_border

        worksheet[f"G{idx}"] = order.status
        worksheet[f"G{idx}"].alignment = align
        worksheet[f"G{idx}"].border = thin_border

        worksheet[f"H{idx}"] = order.status
        worksheet[f"H{idx}"].alignment = align
        worksheet[f"H{idx}"].border = thin_border


    # Save the modified workbook as an Excel file
    workbook.save(os.path.join(report_dir, "order_report.xlsx"))
    #
    # Convert the Excel file to PDF using Aspose.Cells API
    workbook = Workbook(os.path.join(report_dir, "order_report.xlsx"))
    save_options = PdfSaveOptions()
    save_options.setOnePagePerSheet(True)
    workbook.save(os.path.join(report_dir, "order.pdf"), save_options)



def test_get_all_order(db: Session, level, user_id, start_time, end_time):
    pass
