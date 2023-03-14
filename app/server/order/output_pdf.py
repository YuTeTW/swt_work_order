import os

import pywintypes
from openpyxl.styles import Alignment
from sqlalchemy import and_, extract
from sqlalchemy.orm import aliased

from app.models.domain.order import Order
from app.models.domain.order_issue import OrderIssue
from app.models.domain.user import User
from app.models.domain.user_mark_order import UserMarkOrder

# import jpype
# import asposecells


def get_order_db(db, user_id, month):
    # Create aliases for joined tables
    engineer_alias = aliased(User)
    reporter_alias = aliased(User)

    # Specify the primary table for the query using .select_from()
    order_db = db.query(Order, engineer_alias.name, reporter_alias.name, OrderIssue.name,
                        UserMarkOrder.mark).select_from(Order)

    # Join with the aliased tables
    order_db = order_db.outerjoin(
        engineer_alias, Order.engineer_id == engineer_alias.id
    ).outerjoin(
        reporter_alias, Order.reporter_id == reporter_alias.id
    ).outerjoin(
        OrderIssue, Order.order_issue_id == OrderIssue.id
    ).outerjoin(
        UserMarkOrder, and_(
            UserMarkOrder.order_id == Order.id,
            UserMarkOrder.user_id == user_id
        )
    )

    order_db = order_db.filter(extract('month', Order.report_time) == month)

    return order_db.all()


def edit_excel(worksheet, order_db):

    # Create a list of weekday names to use in the report
    week_name = ["一", "二", "三", "四", "五", "六", "日"]

    # Set alignment for the cells
    align = worksheet.Range("A1").HorizontalAlignment

    # Set border
    thin_border = 1  # 1 = thin border

    # 初始化上一行日期
    prev_date = None
    start_row = 4

    def merge_cells(column, start_row, end_row, value):
        merge_range = worksheet.Range(f"{column}{start_row}:{column}{end_row}")
        merge_range.Merge()
        merge_range.Value = value

    def cell_add_value(column, row, value):
        worksheet.Cells(row, column).Value = value
        worksheet.Cells(row, column).HorizontalAlignment = align
        worksheet.Cells(row, column).Borders(9).LineStyle = thin_border


    # Iterate over the query results and populate the worksheet
    for idx, (order, engineer_name, reporter_name, issue_name, mark) in enumerate(order_db, start=4):
        curr_date = order.report_time.date()
        curr_date = pywintypes.Time(curr_date)

        # Merge cells with the same date
        if curr_date != prev_date:
            # Get the end row for the merged cells
            end_row = idx - 1

            # Merge cells and write the date and week
            if start_row < end_row:
                merge_cells("B", start_row, end_row, curr_date)  # 日期
                merge_cells("C", start_row, end_row, week_name[prev_date.weekday()])  # 星期

            # write line
            worksheet.Range(f"B{end_row}:C{end_row}").Borders(9).LineStyle = thin_border

            # Update the start row for the next merged cells
            start_row = idx

            # Update the previous date
            prev_date = curr_date

            # align
            worksheet.Range(f"B{start_row}:C{start_row}").HorizontalAlignment = align

        # Check if this is the last row
        if idx == len(list(order_db)) + 3:
            end_row = idx

            merge_cells("B", start_row, end_row, curr_date)  # 日期
            merge_cells("C", start_row, end_row, week_name[prev_date.weekday()])  # 星期

            # write line
            worksheet.Range(f"B{end_row}:C{end_row}").Borders(9).LineStyle = thin_border

        # write value
        order_status_name = {
            0: "未處理",
            1: "處理中",
            2: "已處理"
        }
        cell_add_value(4, idx, reporter_name)  # 提報人
        cell_add_value(5, idx, issue_name)  # 項目
        cell_add_value(6, idx, "1")  # 處理方式
        cell_add_value(7, idx, order_status_name[order.status])  # 程序


async def get_report(db, user_id, month):
    import win32com.client as win32
    import os

    # 設置Excel和PDF文件的路徑
    excel_path = os.getcwd() + r'\order.xlsx'
    pdf_path = os.getcwd() + r'\file.pdf'
    output_path = os.getcwd() + r'\file2.xlsx'

    # 建立Excel和PDF應用程序實例
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    pdf = win32.gencache.EnsureDispatch('Word.Application')

    workbook = excel.Workbooks.Open(excel_path)

    # 打開Excel文件
    # try:

    worksheet = workbook.Worksheets(1)
    #
    # worksheet.merge()
    order_db = get_order_db(db, user_id, month)

    edit_excel(worksheet, order_db)
    # 將工作表轉換為PDF
    workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
    workbook.SaveAs(output_path)
    # except:
    #     workbook.Close(False)
    #     excel.Quit()
    #     pdf.Quit()

    # 關閉Excel和PDF應用程序
    workbook.Close(False)
    excel.Quit()
    pdf.Quit()



    # import openpyxl
    # from asposecells.api import Workbook, FileFormatType, PdfSaveOptions
    #
    #
    # order_db = get_order_db(db, user_id)
    #
    # # Create directory for report files if it doesn't exist
    # report_dir = os.path.join(os.getcwd(), "db_image/order/report/")
    # os.makedirs(report_dir, exist_ok=True)
    #
    # # Load the Excel workbook
    # workbook = openpyxl.load_workbook('order.xlsx')
    #
    # # Select the worksheet to edit
    # worksheet = workbook['EFAY - 2023 JAN']
    #
    # # write data into excel
    # edit_excel(worksheet, order_db)
    #
    #
    # # Save the modified workbook as an Excel file
    # workbook.save(os.path.join(report_dir, "order_report.xlsx"))
    # #
    # # Convert the Excel file to PDF using Aspose.Cells API
    # workbook = Workbook(os.path.join(report_dir, "order_report.xlsx"))
    # save_options = PdfSaveOptions()
    # save_options.setOnePagePerSheet(True)
    # workbook.save(os.path.join(report_dir, "order.pdf"), save_options)
    # import win32com.client as win32
    # import os
    #
    #
    # order_db = get_order_db(db, user_id)
    #
    # # Create directory for report files if it doesn't exist
    # report_dir = os.path.join(os.getcwd(), "db_image/order/report/")
    # os.makedirs(report_dir, exist_ok=True)
    #
    # # 設置Excel和PDF文件的路徑
    # excel_path = os.getcwd() + r'\order.xlsx'
    # pdf_path = os.getcwd() + r'\order.pdf'
    #
    # # 建立Excel和PDF應用程序實例
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    # pdf = win32.gencache.EnsureDispatch('Word.Application')
    #
    #
    #
    # # 打開Excel文件
    # workbook = excel.Workbooks.Open(excel_path)
    #
    # worksheet = workbook.Worksheets(1)
    #
    # # # write data into excel
    # edit_excel(worksheet, order_db)
    #
    # # 將工作表轉換為PDF
    # workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
    #
    # # 關閉Excel和PDF應用程序
    # workbook.Close(False)
    # excel.Quit()
    # pdf.Quit()